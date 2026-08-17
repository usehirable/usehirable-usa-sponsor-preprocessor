"""Build employer-level U.S. visa sponsor data from DOL LCA disclosure files."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Iterable
from urllib.parse import urljoin, urlsplit, urlunsplit

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


DOL_PERFORMANCE_URL = "https://www.dol.gov/agencies/eta/foreign-labor/performance"
SOURCE_NAME = "U.S. DOL LCA Disclosure"
COUNTRY = "United States"
COUNTRY_CODE = "US"
DISCLOSURE_RE = re.compile(
    r"LCA_Disclosure_Data_FY(?P<year>\d{4})_Q(?P<quarter>[1-4])\.xlsx",
    re.IGNORECASE,
)
EXCLUDED_LINK_TEXT = ("appendix", "worksite")
RELEVANT_VISA_CLASSES = {"H-1B", "H-1B1", "E-3"}
RELEVANT_STATUS_PREFIXES = ("CERTIFIED",)
DEFAULT_CHUNK_SIZE = 5000
CHUNK_FILE_RE = re.compile(r"usa-\d{3}\.json")

HEADER_ALIASES = {
    "employer_name": {
        "employername",
        "employerlegalbusinessname",
        "employerlegalname",
        "petitionername",
        "companyname",
    },
    "employer_city": {
        "employercity",
        "employercityname",
        "employerlocationcity",
        "petitionercity",
    },
    "employer_state": {
        "employerstate",
        "employerstateprovince",
        "employerprovince",
        "employerlocationstate",
        "petitionerstate",
    },
    "visa_class": {
        "visaclass",
        "classofadmission",
        "visa",
        "program",
    },
    "soc_title": {
        "soctitle",
        "socname",
        "sococcupationtitle",
        "socjobtitle",
        "jobtitle",
        "jobtitletext",
    },
    "case_status": {
        "casestatus",
        "status",
        "decisionstatus",
        "casefinalstatus",
    },
}

REQUIRED_COLUMNS = ("employer_name", "employer_city", "employer_state")


@dataclass(frozen=True)
class DisclosureFile:
    url: str
    filename: str
    fiscal_year: int
    quarter: int


@dataclass
class EmployerAggregate:
    company_name: str
    city: str
    state: str
    visa_programs: Counter[str] = field(default_factory=Counter)
    job_titles: Counter[str] = field(default_factory=Counter)
    filing_count: int = 0


@dataclass
class ProcessingStats:
    raw_visa_classes: Counter[str] = field(default_factory=Counter)
    kept_visa_programs: Counter[str] = field(default_factory=Counter)
    skipped_visa_classes: Counter[str] = field(default_factory=Counter)
    skipped_statuses: Counter[str] = field(default_factory=Counter)


def normalize_header(value: object) -> str:
    text = normalize_whitespace(value)
    return re.sub(r"[^a-z0-9]", "", text.lower())


def normalize_whitespace(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_company_name(value: object) -> str:
    return normalize_whitespace(value)


def normalize_group_key(*parts: str) -> tuple[str, ...]:
    return tuple(normalize_whitespace(part).casefold() for part in parts)


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "unknown"


def normalize_url(value: str) -> str:
    parts = urlsplit(value)
    normalized_path = re.sub(r"/{2,}", "/", parts.path)
    return urlunsplit((parts.scheme, parts.netloc, normalized_path, parts.query, parts.fragment))


def normalize_visa_program(value: object) -> str:
    visa_class = normalize_whitespace(value).upper()
    if not visa_class:
        return ""
    if visa_class.startswith("H-1B1"):
        return "H-1B1"
    if visa_class.startswith("H-1B"):
        return "H-1B"
    if visa_class.startswith("E-3"):
        return "E-3"
    return visa_class


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build compact U.S. LCA employer sponsor JSON from DOL disclosure data."
    )
    parser.add_argument(
        "--source-page",
        default=DOL_PERFORMANCE_URL,
        help="DOL performance page used for workbook discovery.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        help="Optional local XLSX workbook. Skips DOL discovery and download.",
    )
    parser.add_argument(
        "--source-url",
        help="Source URL to write into output when --workbook is used.",
    )
    parser.add_argument(
        "--download-dir",
        type=Path,
        help="Optional directory for downloaded workbooks.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("dist/usa-sponsors.json"),
        help="Output JSON path.",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=DEFAULT_CHUNK_SIZE,
        help="Number of sponsor records per chunk file.",
    )
    return parser.parse_args()


def discover_latest_disclosure(source_page: str) -> DisclosureFile:
    response = requests.get(source_page, timeout=45)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    candidates: list[DisclosureFile] = []

    for link in soup.find_all("a", href=True):
        href = str(link["href"])
        text = " ".join(link.stripped_strings)
        combined = f"{href} {text}".lower()
        if any(excluded in combined for excluded in EXCLUDED_LINK_TEXT):
            continue

        match = DISCLOSURE_RE.search(href) or DISCLOSURE_RE.search(text)
        if not match:
            continue

        filename = match.group(0)
        candidates.append(
            DisclosureFile(
                url=normalize_url(urljoin(source_page, href)),
                filename=filename,
                fiscal_year=int(match.group("year")),
                quarter=int(match.group("quarter")),
            )
        )

    if not candidates:
        raise RuntimeError(f"No cumulative LCA disclosure workbook found at {source_page}")

    return max(candidates, key=lambda item: (item.fiscal_year, item.quarter))


def download_file(disclosure: DisclosureFile, destination_dir: Path) -> Path:
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / disclosure.filename

    with requests.get(disclosure.url, stream=True, timeout=(15, 180)) as response:
        response.raise_for_status()
        with destination.open("wb") as output:
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    output.write(chunk)

    if destination.stat().st_size == 0:
        raise RuntimeError(f"Downloaded workbook is empty: {destination}")
    return destination


def find_header_row(rows: Iterable[tuple[object, ...]]) -> tuple[int, list[str]]:
    best_row_number = 0
    best_headers: list[str] = []
    best_score = 0

    for row_number, row in enumerate(rows, start=1):
        headers = [normalize_header(cell) for cell in row]
        header_set = set(headers)
        score = sum(
            1
            for aliases in HEADER_ALIASES.values()
            if aliases.intersection(header_set)
        )
        if score > best_score:
            best_row_number = row_number
            best_headers = headers
            best_score = score
        if score >= len(REQUIRED_COLUMNS):
            return row_number, headers
        if row_number >= 30:
            break

    if best_score:
        return best_row_number, best_headers
    raise RuntimeError("Could not locate the workbook header row.")


def resolve_columns(headers: list[str]) -> dict[str, int | None]:
    columns: dict[str, int | None] = {}
    for column_name, aliases in HEADER_ALIASES.items():
        columns[column_name] = next(
            (index for index, header in enumerate(headers) if header in aliases),
            None,
        )

    missing = [name for name in REQUIRED_COLUMNS if columns[name] is None]
    if missing:
        raise RuntimeError(f"Missing required employer columns: {', '.join(missing)}")

    return columns


def get_cell(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return normalize_whitespace(row[index])


def is_relevant_case(visa_program: str, case_status: str) -> bool:
    case_status_upper = case_status.upper()

    if visa_program and visa_program not in RELEVANT_VISA_CLASSES:
        return False
    if case_status_upper and not case_status_upper.startswith(RELEVANT_STATUS_PREFIXES):
        return False
    return True


def get_metadata_from_filename(filename: str) -> tuple[int, int]:
    match = DISCLOSURE_RE.search(filename)
    if not match:
        raise RuntimeError(f"Could not determine fiscal year and quarter from {filename}")
    return int(match.group("year")), int(match.group("quarter"))


def process_workbook(
    workbook_path: Path,
) -> tuple[dict[tuple[str, str, str], EmployerAggregate], ProcessingStats]:
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    stats = ProcessingStats()
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        row_iterator = worksheet.iter_rows(values_only=True)
        header_row_number, headers = find_header_row(row_iterator)
        columns = resolve_columns(headers)

        aggregates: dict[tuple[str, str, str], EmployerAggregate] = {}
        for row_number, row in enumerate(row_iterator, start=header_row_number + 1):
            company_name = normalize_company_name(get_cell(row, columns["employer_name"]))
            city = normalize_whitespace(get_cell(row, columns["employer_city"])).title()
            state = normalize_whitespace(get_cell(row, columns["employer_state"])).upper()
            if not company_name or not city or not state:
                continue

            raw_visa_class = get_cell(row, columns["visa_class"])
            visa_program = normalize_visa_program(raw_visa_class)
            case_status = get_cell(row, columns["case_status"])
            if raw_visa_class:
                stats.raw_visa_classes[normalize_whitespace(raw_visa_class).upper()] += 1
            if visa_program and visa_program not in RELEVANT_VISA_CLASSES:
                stats.skipped_visa_classes[visa_program] += 1
            if case_status and not case_status.upper().startswith(RELEVANT_STATUS_PREFIXES):
                stats.skipped_statuses[case_status.upper()] += 1
            if not is_relevant_case(visa_program, case_status):
                continue

            key = normalize_group_key(company_name, city, state)
            aggregate = aggregates.setdefault(
                key,
                EmployerAggregate(company_name=company_name, city=city, state=state),
            )
            aggregate.filing_count += 1
            if visa_program:
                aggregate.visa_programs[visa_program] += 1
                stats.kept_visa_programs[visa_program] += 1

            soc_title = get_cell(row, columns["soc_title"])
            if soc_title:
                aggregate.job_titles[normalize_whitespace(soc_title).title()] += 1

            if row_number % 25000 == 0:
                print(f"Processed {row_number:,} workbook rows...", file=sys.stderr)
    finally:
        workbook.close()

    if not aggregates:
        raise RuntimeError("No employer-level sponsor records were produced.")
    return aggregates, stats


def print_processing_stats(stats: ProcessingStats) -> None:
    def format_counter(counter: Counter[str]) -> str:
        if not counter:
            return "none"
        return ", ".join(f"{key}={value:,}" for key, value in counter.most_common(10))

    print(f"Raw visa classes: {format_counter(stats.raw_visa_classes)}", file=sys.stderr)
    print(f"Kept visa programs: {format_counter(stats.kept_visa_programs)}", file=sys.stderr)
    if stats.skipped_visa_classes:
        print(
            f"Skipped visa classes: {format_counter(stats.skipped_visa_classes)}",
            file=sys.stderr,
        )


def build_company_records(
    aggregates: dict[tuple[str, str, str], EmployerAggregate],
    source_url: str,
) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for aggregate in aggregates.values():
        location = f"{aggregate.city}, {aggregate.state}"
        visa_programs = (
            " / ".join(sorted(aggregate.visa_programs))
            if aggregate.visa_programs
            else "H-1B / H-1B1 / E-3"
        )
        top_titles = [title for title, _count in aggregate.job_titles.most_common(5)]

        records.append(
            {
                "company_name": aggregate.company_name,
                "company_slug": slugify(aggregate.company_name),
                "country": COUNTRY,
                "country_code": COUNTRY_CODE,
                "country_slug": "united-states",
                "industry": "",
                "location": location,
                "location_slug": slugify(location),
                "tags": ", ".join(top_titles),
                "sponsor_type": "LCA Employer",
                "visa_program": visa_programs,
                "top_job_titles": top_titles,
                "source": SOURCE_NAME,
                "source_url": source_url,
                "filing_count": aggregate.filing_count,
            }
        )

    return sorted(
        records,
        key=lambda record: (
            str(record["company_name"]).casefold(),
            str(record["location"]).casefold(),
        ),
    )


def get_generated_at() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def get_total_filings(companies: list[dict[str, object]]) -> int:
    return sum(int(company.get("filing_count") or 0) for company in companies)


def build_full_payload(
    companies: list[dict[str, object]],
    source_url: str,
    fiscal_year: int,
    quarter: int,
    version: str,
    generated_at: str,
) -> dict[str, object]:
    return {
        "version": version,
        "generatedAt": generated_at,
        "source": SOURCE_NAME,
        "sourceUrl": source_url,
        "fiscalYear": fiscal_year,
        "quarter": quarter,
        "totalCompanies": len(companies),
        "totalFilings": get_total_filings(companies),
        "companies": companies,
    }


def build_chunk_payloads(
    companies: list[dict[str, object]],
    fiscal_year: int,
    quarter: int,
    version: str,
    chunk_size: int,
) -> list[tuple[str, dict[str, object]]]:
    chunk_count = (len(companies) + chunk_size - 1) // chunk_size
    payloads: list[tuple[str, dict[str, object]]] = []

    for index in range(chunk_count):
        chunk_companies = companies[index * chunk_size : (index + 1) * chunk_size]
        filename = f"usa-{index + 1:03d}.json"
        payloads.append(
            (
                filename,
                {
                    "version": version,
                    "country": COUNTRY,
                    "countryCode": COUNTRY_CODE,
                    "fiscalYear": fiscal_year,
                    "quarter": quarter,
                    "chunk": index + 1,
                    "chunkCount": chunk_count,
                    "chunkSize": chunk_size,
                    "recordCount": len(chunk_companies),
                    "companies": chunk_companies,
                },
            )
        )

    return payloads


def build_manifest_payload(
    full_payload: dict[str, object],
    chunk_payloads: list[tuple[str, dict[str, object]]],
    chunk_size: int,
) -> dict[str, object]:
    chunks = [
        {
            "index": payload["chunk"],
            "file": filename,
            "recordCount": payload["recordCount"],
        }
        for filename, payload in chunk_payloads
    ]
    return {
        "version": full_payload["version"],
        "generatedAt": full_payload["generatedAt"],
        "country": COUNTRY,
        "countryCode": COUNTRY_CODE,
        "source": full_payload["source"],
        "sourceUrl": full_payload["sourceUrl"],
        "fiscalYear": full_payload["fiscalYear"],
        "quarter": full_payload["quarter"],
        "totalCompanies": full_payload["totalCompanies"],
        "totalFilings": full_payload["totalFilings"],
        "chunkSize": chunk_size,
        "chunkCount": len(chunks),
        "chunks": chunks,
    }


def validate_output_contract(
    companies: list[dict[str, object]],
    full_payload: dict[str, object],
    manifest_payload: dict[str, object],
    chunk_payloads: list[tuple[str, dict[str, object]]],
    chunk_size: int,
    source_url: str,
    fiscal_year: int,
    quarter: int,
) -> None:
    if not companies:
        raise RuntimeError("Cannot write USA sponsor output: companies array is empty.")
    if full_payload["totalCompanies"] == 0:
        raise RuntimeError("Cannot write USA sponsor output: totalCompanies is 0.")
    if not source_url:
        raise RuntimeError("Cannot write USA sponsor output: sourceUrl is empty.")
    if not fiscal_year:
        raise RuntimeError("Cannot write USA sponsor output: fiscalYear is missing.")
    if not quarter:
        raise RuntimeError("Cannot write USA sponsor output: quarter is missing.")

    for index, company in enumerate(companies):
        if not company.get("company_name"):
            raise RuntimeError(f"Cannot write USA sponsor output: company_name missing at index {index}.")

    filenames = [filename for filename, _payload in chunk_payloads]
    if len(filenames) != len(set(filenames)):
        raise RuntimeError("Cannot write USA sponsor output: duplicate chunk filenames.")

    combined_companies: list[dict[str, object]] = []
    for index, (_filename, payload) in enumerate(chunk_payloads):
        record_count = int(payload["recordCount"])
        if record_count == 0:
            raise RuntimeError("Cannot write USA sponsor output: chunk contains 0 records.")
        if index < len(chunk_payloads) - 1 and record_count != chunk_size:
            raise RuntimeError(
                f"Cannot write USA sponsor output: chunk {index + 1} has {record_count} records, expected {chunk_size}."
            )
        combined_companies.extend(payload["companies"])

    if sum(chunk["recordCount"] for chunk in manifest_payload["chunks"]) != full_payload["totalCompanies"]:
        raise RuntimeError("Cannot write USA sponsor output: chunk record totals do not match totalCompanies.")
    if manifest_payload["chunkCount"] != len(manifest_payload["chunks"]):
        raise RuntimeError("Cannot write USA sponsor output: manifest chunkCount does not match chunks length.")
    if combined_companies != companies:
        raise RuntimeError("Cannot write USA sponsor output: chunked companies do not match full dataset order.")


def write_json(path: Path, payload: dict[str, object]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )


def remove_stale_chunk_files(output_dir: Path) -> None:
    for path in output_dir.glob("usa-*.json"):
        if CHUNK_FILE_RE.fullmatch(path.name):
            path.unlink()


def write_outputs(
    output_path: Path,
    companies: list[dict[str, object]],
    source_url: str,
    fiscal_year: int,
    quarter: int,
    chunk_size: int,
) -> dict[str, object]:
    if chunk_size <= 0:
        raise RuntimeError("Chunk size must be greater than 0.")

    version = date.today().isoformat()
    generated_at = get_generated_at()
    output_dir = output_path.parent
    full_payload = build_full_payload(
        companies=companies,
        source_url=source_url,
        fiscal_year=fiscal_year,
        quarter=quarter,
        version=version,
        generated_at=generated_at,
    )
    chunk_payloads = build_chunk_payloads(companies, fiscal_year, quarter, version, chunk_size)
    manifest_payload = build_manifest_payload(full_payload, chunk_payloads, chunk_size)
    validate_output_contract(
        companies=companies,
        full_payload=full_payload,
        manifest_payload=manifest_payload,
        chunk_payloads=chunk_payloads,
        chunk_size=chunk_size,
        source_url=source_url,
        fiscal_year=fiscal_year,
        quarter=quarter,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(output_path, full_payload)
    remove_stale_chunk_files(output_dir)
    for filename, payload in chunk_payloads:
        write_json(output_dir / filename, payload)
    write_json(output_dir / "manifest.json", manifest_payload)
    return manifest_payload


def run() -> None:
    args = parse_args()

    if args.workbook:
        workbook_path = args.workbook
        source_url = normalize_url(args.source_url) if args.source_url else str(workbook_path)
        fiscal_year, quarter = get_metadata_from_filename(workbook_path.name)
    else:
        disclosure = discover_latest_disclosure(args.source_page)
        source_url = disclosure.url
        fiscal_year = disclosure.fiscal_year
        quarter = disclosure.quarter
        if args.download_dir:
            workbook_path = download_file(disclosure, args.download_dir)
        else:
            with TemporaryDirectory() as temp_dir:
                workbook_path = download_file(disclosure, Path(temp_dir))
                aggregates, stats = process_workbook(workbook_path)
                print_processing_stats(stats)
                companies = build_company_records(aggregates, source_url)
                manifest = write_outputs(args.output, companies, source_url, fiscal_year, quarter, args.chunk_size)
                print(
                    f"Wrote {len(companies):,} companies across {manifest['chunkCount']} chunks to {args.output.parent}"
                )
                return

    aggregates, stats = process_workbook(workbook_path)
    print_processing_stats(stats)
    companies = build_company_records(aggregates, source_url)
    manifest = write_outputs(args.output, companies, source_url, fiscal_year, quarter, args.chunk_size)
    print(
        f"Wrote {len(companies):,} companies across {manifest['chunkCount']} chunks to {args.output.parent}"
    )


def main() -> int:
    try:
        run()
    except Exception as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
